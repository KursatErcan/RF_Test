from openpyxl import load_workbook
import serial, time


file = r"RemoteControl.xlsx"
workbook=load_workbook(file)

serial_ = serial.Serial(port="COM47",
                        baudrate=9600,
                        bytesize=serial.EIGHTBITS,
                        parity=serial.PARITY_NONE,
                        stopbits=serial.STOPBITS_ONE,
                        timeout=1)

time.sleep(2.0)

def read(book, sheet, row, column):
    try:
        workbook = load_workbook(book)
        worksheet = workbook.get_sheet_by_name(sheet)
        result = worksheet.cell(row=row, column=column).value
        workbook.close()
        print(f'-------------> Excel Read ---> Sheet:{sheet} Line:{row} Column:{column} ---> Data:{result}')
        return result
    except Exception as error:
        print(f'####### ERROR !!! ####### -----> Excel Read: {error}')
        return None

def get_last_line(book, sheet):
    try:
        workbook = load_workbook(book)
        worksheet = workbook.get_sheet_by_name(sheet)
        result = worksheet.max_row
        workbook.close()
        # print(f'-------------> Excel Get Last Line from Sheet:{sheet} ---> Last Line:{result}')
        return result
    except Exception as error:
        print(f'####### ERROR !!! ####### -----> Excel Get Last Line: {error}')
        return None


for sheet in workbook.get_sheet_names():

    for line in range(int(float(get_last_line(file, sheet)))):
        buttonName = read(file, sheet, (line+1), 1)
        signal = read(file, sheet, (line+1), 2)
        bit = read(file, sheet, (line+1), 3)
        sendValue = "{};{};{}\n".format(int(signal, 16), int(bit), buttonName)
        print(serial_.read(1024).decode("utf-8"))
        serial_.write(sendValue.encode("utf-8"))
        time.sleep(1.0)
        serial_.flushInput()
        serial_.flushOutput()

serial_.close()

